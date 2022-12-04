import openpyxl

def getRowCount(filepath,sheetname):
    workbook=openpyxl.load_workbook(filepath)
    sheet=workbook[sheetname]
    row=sheet.max_row
    return row

def getColumnCount(filepath,sheetname):
    workbook=openpyxl.load_workbook(filepath)
    sheet=workbook[sheetname]
    column=sheet.max_column
    return column

def getEXlData(filepath,sheetname,rowcount,columncount):
    workbook=openpyxl.load_workbook(filepath)
    sheet=workbook[sheetname]
    return sheet.cell(row=rowcount,column=columncount).value

def getWRTData(filepath,sheetname,rowcount,columncount,data):
    workbook=openpyxl.load_workbook(filepath)
    sheet=workbook[sheetname]
    sheet.cell(row=rowcount,column=columncount).value= data
    return workbook.save(filepath)